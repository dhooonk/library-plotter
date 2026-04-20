"""
excel_exporter.py
-----------------
분석 결과를 기반으로 엑셀(.xlsx) 통합 문서를 생성하고 저장하는 모듈.
- Sheet 1: Raw 데이터 (Vd 혹은 Vg별로 구조화, 정렬된 셀 서식 작성)
- Sheet 2: 차트 데이터 (matplotlib 차트를 백그라운드용 PNG 고화질로 렌더링 후 삽입)
"""

import io
import os
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from matplotlib.figure import Figure

import output_curve as oc_module
import transfer_curve as tc_module

# ────────────────────────── 스타일 상수 설계 ──────────────────────────
# 엑셀 헤더와 내용에 들어가는 각종 폰트 및 채우기 스타일을 공통 정의합니다.
_HEADER_FILL    = PatternFill("solid", fgColor="1F3864")  # 진한 파랑 (최상단 파라미터 헤더)
_PARAM_FILL     = PatternFill("solid", fgColor="2E4057")  # 미디엄 파랑 (V/A 축 이름 헤더)
_ALT_FILL       = PatternFill("solid", fgColor="F2F7FF")  # 연한 파랑 (데이터 가독성 개선-교차 행)
_HEADER_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_PARAM_FONT     = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_DATA_FONT      = Font(name="Calibri", size=10)
_THIN_BORDER    = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)
_CENTER         = Alignment(horizontal="center", vertical="center")
# 데이터별 출력 표시 형식 (지수형, 소수점 셋째자리 고정형 등)
_NUM_FMT_SCI    = "0.000E+00"
_NUM_FMT_VOLT   = "0.000"


def _apply_cell(ws, row, col, value, fill=None, font=None, alignment=None,
                border=None, number_format=None):
    """지정된 셀 위치(row, col)에 들어갈 데이터와 속성(색, 폰트, 보더)을 한 번에 적용하는 헬퍼 함수"""
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def _figure_to_xl_image(fig: Figure, width_px: int = 680, height_px: int = 480) -> XLImage:
    """
    생성된 matplotlib Figure(차트 객체)를 메모리 버퍼상에 `.png` 형식으로
    작성하고 크기에 맞춰 축척한 openpyxl Image 객체로 반환.
    """
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight",
                facecolor=fig.get_facecolor(), dpi=130)
    buf.seek(0)
    img = XLImage(buf)
    # 셀 크기 조정을 위한 픽셀 리사이징 매핑
    img.width = width_px
    img.height = height_px
    return img


# ───────────────────────── Transfer Curve 내보내기 ─────────────────────────

def export_transfer_curve(
    grouped_data: dict,
    fig: Figure,
    save_path: str,
    log_scale: bool = True,
    xlim: tuple = None,
    ylim: tuple = None
):
    """
    Transfer Curve 결과를 엑셀 파일(저장 경로 지정)로 저장합니다. 
    1st 트랙: Vd별 두 가지 열(Vgs, Id) 로 원시 데이터 생성
    """
    wb = Workbook()

    # ── Sheet 1: Raw 데이터 작성 부분 ──
    ws_data = wb.active
    ws_data.title = "Raw Data"

    sorted_vds = sorted(grouped_data.keys())

    # 기본 행 폭 설정
    ws_data.row_dimensions[1].height = 20
    ws_data.row_dimensions[2].height = 18

    # 헤더 행 데이터 삽입
    col = 1
    vd_col_starts = {}
    for vd in sorted_vds:
        vd_col_starts[vd] = col
        vd_label = f"Vd = {_fmt(vd)} V"
        
        # 병합 헤더로 하나의 Vd 단위 라벨링
        ws_data.merge_cells(
            start_row=1, start_column=col,
            end_row=1, end_column=col + 1
        )
        _apply_cell(ws_data, 1, col, vd_label,
                    fill=_HEADER_FILL, font=_HEADER_FONT,
                    alignment=_CENTER, border=_THIN_BORDER)
                    
        # 서브 헤더 (Vgs, Id 열 부여)
        _apply_cell(ws_data, 2, col, "Vgs (V)",
                    fill=_PARAM_FILL, font=_PARAM_FONT,
                    alignment=_CENTER, border=_THIN_BORDER)
        _apply_cell(ws_data, 2, col + 1, "Id (A)",
                    fill=_PARAM_FILL, font=_PARAM_FONT,
                    alignment=_CENTER, border=_THIN_BORDER)
                    
        # 컬럼 너비 재조정
        ws_data.column_dimensions[get_column_letter(col)].width = 14
        ws_data.column_dimensions[get_column_letter(col + 1)].width = 16
        col += 2

    # 해당 데이터 쌍 그룹 중 데이터 개수가 제일 많은 row 수를 감지.
    max_len = max(len(d["x"]) for d in grouped_data.values()) if grouped_data else 0
    
    # 데이터 행 삽입 루프 (개별 데이터별 x, y 기입)
    for row_i in range(max_len):
        excel_row = row_i + 3
        # 홀수 줄에만 연한 배경색상을 먹이기 위한 이분탐색(홀짝 체크)
        is_odd = (row_i % 2 == 0)
        
        for vd in sorted_vds:
            c = vd_col_starts[vd]
            x_arr = grouped_data[vd]["x"]
            y_arr = grouped_data[vd]["y"]
            row_fill = None if is_odd else _ALT_FILL
            
            # 배열 접근이 가능한 유효 범위인 경우 값 주입
            if row_i < len(x_arr):
                _apply_cell(ws_data, excel_row, c, float(x_arr[row_i]),
                            fill=row_fill, font=_DATA_FONT,
                            alignment=_CENTER, border=_THIN_BORDER,
                            number_format=_NUM_FMT_VOLT)
                _apply_cell(ws_data, excel_row, c + 1, float(y_arr[row_i]),
                            fill=row_fill, font=_DATA_FONT,
                            alignment=_CENTER, border=_THIN_BORDER,
                            number_format=_NUM_FMT_SCI)
            # 데이터 개수가 달라 빈칸인 경우는 공백 렌더링 유지
            else:
                _apply_cell(ws_data, excel_row, c, None,
                            fill=row_fill, border=_THIN_BORDER)
                _apply_cell(ws_data, excel_row, c + 1, None,
                            fill=row_fill, border=_THIN_BORDER)

    # 3행 밑으로 스크롤 고정되도록 A3 위치 고정 (틀 고정)
    ws_data.freeze_panes = "A3"

    # ── Sheet 2: 차트 엑셀 내 삽입 부분 ──
    ws_chart = wb.create_sheet("Transfer Curve Chart")
    ws_chart.sheet_view.showGridLines = False
    ws_chart.column_dimensions["A"].width = 1

    title_font = Font(name="Calibri", bold=True, size=13, color="1F3864")
    _apply_cell(ws_chart, 1, 2, "TFT Transfer Curve (Vgs-Id)",
                font=title_font, alignment=_CENTER)
    _apply_cell(ws_chart, 2, 2,
                f"Y축: {'로그 스케일 (Log)' if log_scale else '선형 스케일 (Linear)'}",
                font=Font(name="Calibri", italic=True, size=10, color="555555"),
                alignment=_CENTER)

    # UI의 Tkinter fig 인스턴스를 바로 추출하면 종속성 문제/화면 깨짐/빈 도화지 상태가 생길 수 있으므로,
    # Agg 백엔드로 독립적으로 차트 객체를 새롭게 생성하여 버퍼링 후 이미지로 내보냅니다.
    export_fig = tc_module.create_transfer_figure(
        grouped_data, log_scale=log_scale,
        title="TFT Transfer Curve  (Vgs - Id)",
        xlim=xlim, ylim=ylim
    )
    xl_img = _figure_to_xl_image(export_fig, width_px=700, height_px=500)
    plt.close(export_fig)
    ws_chart.add_image(xl_img, "B4")

    # '.xlsx' 형태로 저장 강제화
    if not save_path.endswith(".xlsx"):
        save_path += ".xlsx"
    wb.save(save_path)


# ───────────────────────── Output Curve 내보내기 ─────────────────────────

def export_output_curve(
    grouped_data: dict,
    fig: Figure,
    save_path: str,
    log_scale: bool = False,
    xlim: tuple = None,
    ylim: tuple = None
):
    """
    Output Curve 결과를 엑셀로 저장. 동작 원리는 `export_transfer_curve`와 완벽히 동일하며, 
    그룹화 파라미터가 "Vg" 기준으로 작동한다는 차이가 존재.
    """
    wb = Workbook()

    ws_data = wb.active
    ws_data.title = "Raw Data"

    sorted_vgs = sorted(grouped_data.keys())

    ws_data.row_dimensions[1].height = 20
    ws_data.row_dimensions[2].height = 18

    col = 1
    vg_col_starts = {}
    for vg in sorted_vgs:
        vg_col_starts[vg] = col
        vg_label = f"Vg = {_fmt(vg)} V"
        ws_data.merge_cells(
            start_row=1, start_column=col,
            end_row=1, end_column=col + 1
        )
        _apply_cell(ws_data, 1, col, vg_label,
                    fill=_HEADER_FILL, font=_HEADER_FONT,
                    alignment=_CENTER, border=_THIN_BORDER)
        _apply_cell(ws_data, 2, col, "Vds (V)",
                    fill=_PARAM_FILL, font=_PARAM_FONT,
                    alignment=_CENTER, border=_THIN_BORDER)
        _apply_cell(ws_data, 2, col + 1, "Id (A)",
                    fill=_PARAM_FILL, font=_PARAM_FONT,
                    alignment=_CENTER, border=_THIN_BORDER)
        ws_data.column_dimensions[get_column_letter(col)].width = 14
        ws_data.column_dimensions[get_column_letter(col + 1)].width = 16
        col += 2

    max_len = max(len(d["x"]) for d in grouped_data.values()) if grouped_data else 0
    for row_i in range(max_len):
        excel_row = row_i + 3
        is_odd = (row_i % 2 == 0)
        for vg in sorted_vgs:
            c = vg_col_starts[vg]
            x_arr = grouped_data[vg]["x"]
            y_arr = grouped_data[vg]["y"]
            row_fill = None if is_odd else _ALT_FILL
            if row_i < len(x_arr):
                _apply_cell(ws_data, excel_row, c, float(x_arr[row_i]),
                            fill=row_fill, font=_DATA_FONT,
                            alignment=_CENTER, border=_THIN_BORDER,
                            number_format=_NUM_FMT_VOLT)
                _apply_cell(ws_data, excel_row, c + 1, float(y_arr[row_i]),
                            fill=row_fill, font=_DATA_FONT,
                            alignment=_CENTER, border=_THIN_BORDER,
                            number_format=_NUM_FMT_SCI)
            else:
                _apply_cell(ws_data, excel_row, c, None,
                            fill=row_fill, border=_THIN_BORDER)
                _apply_cell(ws_data, excel_row, c + 1, None,
                            fill=row_fill, border=_THIN_BORDER)

    ws_data.freeze_panes = "A3"

    ws_chart = wb.create_sheet("Output Curve Chart")
    ws_chart.sheet_view.showGridLines = False
    ws_chart.column_dimensions["A"].width = 1

    title_font = Font(name="Calibri", bold=True, size=13, color="1F3864")
    _apply_cell(ws_chart, 1, 2, "TFT Output Curve (Vd-Id)",
                font=title_font, alignment=_CENTER)
    _apply_cell(ws_chart, 2, 2,
                f"Y축: {'로그 스케일 (Log)' if log_scale else '선형 스케일 (Linear)'}",
                font=Font(name="Calibri", italic=True, size=10, color="555555"),
                alignment=_CENTER)

    # 출력 전용 figure 생성으로 빈 Blank 도화지(UI 객체 종속 오류) 렌더링 방어
    export_fig = oc_module.create_output_figure(
        grouped_data, log_scale=log_scale,
        title="TFT Output Curve  (Vd - Id)",
        xlim=xlim, ylim=ylim
    )
    xl_img = _figure_to_xl_image(export_fig, width_px=700, height_px=500)
    plt.close(export_fig)
    ws_chart.add_image(xl_img, "B4")

    if not save_path.endswith(".xlsx"):
        save_path += ".xlsx"
    wb.save(save_path)


def _fmt(v: float) -> str:
    """Float 소수점 정리 내부 헬퍼 (vd=1.0 인경우 문자열 1 반환)"""
    if v == int(v):
        return str(int(v))
    return f"{v:.3g}"
