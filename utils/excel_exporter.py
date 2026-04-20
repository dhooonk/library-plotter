import io
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from matplotlib.figure import Figure

import core.output_curve as oc_module
import core.transfer_curve as tc_module

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_PARAM_FILL  = PatternFill("solid", fgColor="2E4057")
_ALT_FILL    = PatternFill("solid", fgColor="F2F7FF")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_PARAM_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_DATA_FONT   = Font(name="Calibri", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)
_CENTER      = Alignment(horizontal="center", vertical="center")
_NUM_FMT_SCI  = "0.000E+00"
_NUM_FMT_VOLT = "0.000"


def _apply_cell(ws, row, col, value, fill=None, font=None, alignment=None,
                border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def _figure_to_xl_image(fig: Figure, width_px: int = 680, height_px: int = 480) -> XLImage:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight",
                facecolor=fig.get_facecolor(), dpi=130)
    buf.seek(0)
    img = XLImage(buf)
    img.width = width_px
    img.height = height_px
    return img


def _write_raw_data_sheet(ws, grouped_data: dict, param_name: str, x_header: str):
    """파라미터별로 묶인 Raw Data를 2열(x, y) 블록 구조로 워크시트에 기록."""
    sorted_params = sorted(grouped_data.keys())
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18

    col = 1
    param_col_starts = {}
    for param_val in sorted_params:
        param_col_starts[param_val] = col
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        _apply_cell(ws, 1, col, f"{param_name} = {_fmt(param_val)} V",
                    fill=_HEADER_FILL, font=_HEADER_FONT, alignment=_CENTER, border=_THIN_BORDER)
        _apply_cell(ws, 2, col, x_header,
                    fill=_PARAM_FILL, font=_PARAM_FONT, alignment=_CENTER, border=_THIN_BORDER)
        _apply_cell(ws, 2, col + 1, "Id (A)",
                    fill=_PARAM_FILL, font=_PARAM_FONT, alignment=_CENTER, border=_THIN_BORDER)
        ws.column_dimensions[get_column_letter(col)].width = 14
        ws.column_dimensions[get_column_letter(col + 1)].width = 16
        col += 2

    max_len = max(len(d["x"]) for d in grouped_data.values()) if grouped_data else 0
    for row_i in range(max_len):
        excel_row = row_i + 3
        row_fill = None if row_i % 2 == 0 else _ALT_FILL
        for param_val in sorted_params:
            c = param_col_starts[param_val]
            x_arr = grouped_data[param_val]["x"]
            y_arr = grouped_data[param_val]["y"]
            if row_i < len(x_arr):
                _apply_cell(ws, excel_row, c, float(x_arr[row_i]),
                            fill=row_fill, font=_DATA_FONT, alignment=_CENTER,
                            border=_THIN_BORDER, number_format=_NUM_FMT_VOLT)
                _apply_cell(ws, excel_row, c + 1, float(y_arr[row_i]),
                            fill=row_fill, font=_DATA_FONT, alignment=_CENTER,
                            border=_THIN_BORDER, number_format=_NUM_FMT_SCI)
            else:
                _apply_cell(ws, excel_row, c, None, fill=row_fill, border=_THIN_BORDER)
                _apply_cell(ws, excel_row, c + 1, None, fill=row_fill, border=_THIN_BORDER)

    ws.freeze_panes = "A3"


def export_transfer_curve(
    grouped_data: dict,
    fig: Figure,
    save_path: str,
    log_scale: bool = True,
    xlim: tuple = None,
    ylim: tuple = None,
):
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Raw Data"
    _write_raw_data_sheet(ws_data, grouped_data, param_name="Vd", x_header="Vgs (V)")

    ws_chart = wb.create_sheet("Transfer Curve Chart")
    ws_chart.sheet_view.showGridLines = False
    ws_chart.column_dimensions["A"].width = 1
    _apply_cell(ws_chart, 1, 2, "TFT Transfer Curve (Vgs-Id)",
                font=Font(name="Calibri", bold=True, size=13, color="1F3864"),
                alignment=_CENTER)
    _apply_cell(ws_chart, 2, 2,
                f"Y축: {'로그 스케일 (Log)' if log_scale else '선형 스케일 (Linear)'}",
                font=Font(name="Calibri", italic=True, size=10, color="555555"),
                alignment=_CENTER)

    # UI의 Tkinter Figure를 재사용하면 백엔드 충돌이 발생하므로 독립 Figure로 재생성
    export_fig = tc_module.create_transfer_figure(
        grouped_data, log_scale=log_scale,
        title="TFT Transfer Curve  (Vgs - Id)",
        xlim=xlim, ylim=ylim,
    )
    ws_chart.add_image(_figure_to_xl_image(export_fig, width_px=700, height_px=500), "B4")
    plt.close(export_fig)

    if not save_path.endswith(".xlsx"):
        save_path += ".xlsx"
    wb.save(save_path)


def export_output_curve(
    grouped_data: dict,
    fig: Figure,
    save_path: str,
    log_scale: bool = False,
    xlim: tuple = None,
    ylim: tuple = None,
):
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Raw Data"
    _write_raw_data_sheet(ws_data, grouped_data, param_name="Vg", x_header="Vds (V)")

    ws_chart = wb.create_sheet("Output Curve Chart")
    ws_chart.sheet_view.showGridLines = False
    ws_chart.column_dimensions["A"].width = 1
    _apply_cell(ws_chart, 1, 2, "TFT Output Curve (Vd-Id)",
                font=Font(name="Calibri", bold=True, size=13, color="1F3864"),
                alignment=_CENTER)
    _apply_cell(ws_chart, 2, 2,
                f"Y축: {'로그 스케일 (Log)' if log_scale else '선형 스케일 (Linear)'}",
                font=Font(name="Calibri", italic=True, size=10, color="555555"),
                alignment=_CENTER)

    # UI의 Tkinter Figure를 재사용하면 백엔드 충돌이 발생하므로 독립 Figure로 재생성
    export_fig = oc_module.create_output_figure(
        grouped_data, log_scale=log_scale,
        title="TFT Output Curve  (Vd - Id)",
        xlim=xlim, ylim=ylim,
    )
    ws_chart.add_image(_figure_to_xl_image(export_fig, width_px=700, height_px=500), "B4")
    plt.close(export_fig)

    if not save_path.endswith(".xlsx"):
        save_path += ".xlsx"
    wb.save(save_path)


def _fmt(v: float) -> str:
    if v == int(v):
        return str(int(v))
    return f"{v:.3g}"
